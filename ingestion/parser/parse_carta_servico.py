"""
==============================================================================
        DUQUE IA - INGESTOR: CARTA DE SERVICO MUNICIPAL (Excel)
==============================================================================
Processa o arquivo CARTA_DE_SERVICO_AJUSTE_23.05.26.xlsx e grava os servicos
municipais no banco SQLite com embeddings reais (gemini-embedding-2).

Estrutura da planilha:
  Linha 1-2 : Cabecalho / data de geracao
  Linha 3   : Nomes das colunas reais
  Linha 4+  : Dados dos servicos (ID, Servico, Orgao, Categoria, O que e, Como acessar, etc.)

Cada servico vira 1 chunk no banco.
==============================================================================
"""

import io
import os
import sys
import json
import sqlite3
import time

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, ROOT)

from dotenv import load_dotenv
load_dotenv(os.path.join(ROOT, ".env"))

import openpyxl
from utils.gemini_client import GeminiClient

# ---- Configuracao -----------------------------------------------------------
XLSX_PATH   = os.path.join(ROOT, "data", "knowledge", "CARTA_DE_SERVICO_AJUSTE_23.05.26.xlsx")
DB_PATH     = os.path.join(ROOT, "data", "db", "duque_ia.db")
SOURCE_NAME = "CARTA_DE_SERVICO_23.05.26.xlsx"
CATEGORY    = "carta_servicos"
DELAY_SEC   = 0.15   # pausa entre embeddings para nao estourar rate limit
BATCH_COMMIT = 20    # commit a cada N insercoes

def carregar_planilha(path: str) -> list[dict]:
    """Le o Excel e retorna lista de dicionarios com os dados de cada servico."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Servicos"]

    # A linha 3 contem os cabecalhos reais
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(3, col).value
        headers.append(str(val).strip() if val else f"col_{col}")

    servicos = []
    for row in range(4, ws.max_row + 1):
        row_data = {}
        vazio = True
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row, col_idx).value
            if val:
                row_data[header] = str(val).strip()
                vazio = False
        if not vazio:
            servicos.append(row_data)

    return servicos, headers


def servico_para_texto(servico: dict) -> str:
    """Converte um servico em texto rico para embedding e busca semantica."""
    partes = []

    nome     = servico.get("Serviço", servico.get("Servico", ""))
    orgao    = servico.get("Órgão", servico.get("Orgao", ""))
    cat      = servico.get("Categoria", "")
    o_que_e  = servico.get("O que é o serviço", servico.get("O que e o servico", ""))
    como     = servico.get("Como acessar", "")
    endereco = servico.get("Endereço", servico.get("Endereco", ""))
    doctos   = servico.get("Documentos necessários", servico.get("Documentos necessarios", ""))
    prazo    = servico.get("Prazo", "")
    gratuito = servico.get("Gratuito", "")
    canal    = servico.get("Canal de atendimento", "")

    if nome:
        partes.append(f"Servico: {nome}")
    if orgao:
        partes.append(f"Orgao responsavel: {orgao}")
    if cat:
        partes.append(f"Categoria: {cat}")
    if o_que_e:
        partes.append(f"Descricao: {o_que_e}")
    if como:
        partes.append(f"Como acessar: {como}")
    if endereco:
        partes.append(f"Endereco: {endereco}")
    if doctos:
        partes.append(f"Documentos necessarios: {doctos}")
    if prazo:
        partes.append(f"Prazo: {prazo}")
    if gratuito:
        partes.append(f"Gratuito: {gratuito}")
    if canal:
        partes.append(f"Canal de atendimento: {canal}")

    return " | ".join(partes)


def ingerir_carta_servico():
    print("=" * 70)
    print("   DUQUE IA - INGESTAO: CARTA DE SERVICO MUNICIPAL")
    print("=" * 70)

    # 1. Carrega a planilha
    print(f"\n[1/4] Carregando planilha: {os.path.basename(XLSX_PATH)}")
    servicos, headers = carregar_planilha(XLSX_PATH)
    print(f"      Colunas encontradas : {len(headers)}")
    print(f"      Servicos carregados : {len(servicos)}")

    # Mostra amostra
    if servicos:
        print(f"\n      Amostra do 1o servico:")
        s = servicos[0]
        for k, v in list(s.items())[:6]:
            print(f"        {k}: {v[:60]}")

    # 2. Conecta ao banco
    print(f"\n[2/4] Conectando ao banco de dados: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    # Remove entradas antigas desta fonte (re-ingestao limpa)
    cur.execute("DELETE FROM duque_ia_chunks WHERE source = ?", (SOURCE_NAME,))
    removidos = cur.rowcount
    conn.commit()
    if removidos > 0:
        print(f"      Removidos {removidos} chunks antigos desta fonte.")

    # 3. Inicializa o cliente Gemini
    print(f"\n[3/4] Inicializando cliente Gemini para embeddings...")
    client = GeminiClient()

    # 4. Processa e insere cada servico
    print(f"\n[4/4] Processando {len(servicos)} servicos...\n")
    inseridos = 0
    erros     = 0

    for i, servico in enumerate(servicos):
        servico_id = servico.get("ID", str(i + 1))
        nome       = servico.get("Serviço", servico.get("Servico", f"Servico #{servico_id}"))
        texto      = servico_para_texto(servico)

        try:
            # Gera embedding
            vetor = client.get_embedding(texto[:3000], is_query=False)

            # Metadata JSON
            meta = json.dumps({
                "id"     : servico_id,
                "title"  : nome,
                "orgao"  : servico.get("Órgão", ""),
                "cat"    : servico.get("Categoria", ""),
                "source" : SOURCE_NAME
            }, ensure_ascii=False)

            cur.execute("""
                INSERT INTO duque_ia_chunks (source, category, content, embedding, metadata)
                VALUES (?, ?, ?, ?, ?)
            """, (SOURCE_NAME, CATEGORY, texto, json.dumps(vetor), meta))

            inseridos += 1

            # Commit a cada BATCH_COMMIT insercoes
            if inseridos % BATCH_COMMIT == 0:
                conn.commit()
                print(f"  [{inseridos:>3}/{len(servicos)}] chunks gravados...")

            time.sleep(DELAY_SEC)

        except Exception as e:
            erros += 1
            print(f"  [ERRO] Servico ID {servico_id} ({nome[:40]}): {e}")

    conn.commit()
    conn.close()

    # 5. Resultado final
    print("\n" + "=" * 70)
    print(f"  Concluido!")
    print(f"  Servicos inseridos : {inseridos}")
    print(f"  Erros              : {erros}")
    print(f"  Categoria no banco : {CATEGORY}")
    print(f"  Fonte registrada   : {SOURCE_NAME}")
    print(f"  Modelo de embedding: {client._working_embedding_model}")
    print("=" * 70)

    if inseridos > 0:
        print(f"\n  Para testar, rode:")
        print(f"  python scripts/tests/test_ask.py \"Como fazer alteracao de IPTU?\"")
        print(f"  python scripts/tests/test_ask.py \"Quais servicos da Secretaria de Saude?\"")


if __name__ == "__main__":
    ingerir_carta_servico()
