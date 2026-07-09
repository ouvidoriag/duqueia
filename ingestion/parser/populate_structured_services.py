import os
import sys
import sqlite3
import openpyxl
import re

# Garante acesso à pasta utils e root
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, ROOT)

XLSX_PATH = os.path.join(ROOT, "data", "knowledge", "CARTA_DE_SERVICO_AJUSTE_23.05.26.xlsx")
DB_PATH = os.path.join(ROOT, "agent", "duque_ia.db")

def clean_name(name: str) -> str:
    return name.strip() if name else ""

def generate_sec_code(name: str) -> str:
    """Gera um código abreviado para a secretaria (ex: Secretaria Municipal de Saúde -> SMS)."""
    name_upper = name.upper()
    # Mapeamentos comuns conhecidos
    mappings = {
        "SAÚDE": "SMS",
        "FAZENDA": "SMF",
        "EDUCAÇÃO": "SME",
        "TRANSPORTES": "SMT",
        "OBRAS": "SMO",
        "ADMINISTRAÇÃO": "SMA",
        "MEIO AMBIENTE": "SMMA",
        "ASSISTÊNCIA SOCIAL": "SMASDH",
        "CULTURA": "SMC",
        "URBANISMO": "SMU",
        "PROCURADORIA": "PGM",
        "PREVIDÊNCIA": "IPMDC",
        "FUNDAÇÃO DE APOIO": "FUNDEC"
    }
    for key, val in mappings.items():
        if key in name_upper:
            return val
            
    # Se não mapeado, pega as primeiras letras de cada palavra significativa
    words = [w for w in name_upper.split() if w not in ["DE", "E", "DO", "DA", "PARA", "MUNICIPAL"]]
    if len(words) >= 2:
        return "".join(w[0] for w in words)[:5]
    return name_upper[:4]

def extract_phones(text: str) -> list[str]:
    """Extrai telefones brasileiros do texto."""
    if not text:
        return []
    # Encontra formatos: (21) 99999-9999, (21) 2672-5650, 26725650, 0800...
    pattern = r'(?:\(?\d{2}\)?\s?)?\d{4,5}[-\s]?\d{4}|\b0800\s?\d{3}\s?\d{4}\b'
    matches = re.findall(pattern, text)
    return [m.strip() for m in matches if len(m.strip()) >= 8]

def extract_emails(text: str) -> list[str]:
    """Extrai e-mails do texto."""
    if not text:
        return []
    pattern = r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+'
    return re.findall(pattern, text)

def extract_links(text: str) -> list[str]:
    """Extrai links/URLs do texto."""
    if not text:
        return []
    pattern = r'https?://[^\s,\)\"\']+'
    return re.findall(pattern, text)

def split_steps(text: str) -> list[str]:
    """Divide o campo 'Como acessar' em etapas discretas."""
    if not text:
        return []
    # Divide por quebra de linha ou marcadores de lista numerados/bullets
    lines = re.split(r'\n|(?:\d+[\.\-\)\s]+)|[•\-\*]\s+', text)
    steps = [l.strip() for l in lines if l.strip() and len(l.strip()) > 5]
    return steps if steps else [text.strip()]

def split_documents(text: str) -> list[str]:
    """Divide os documentos necessários em lista de strings."""
    if not text:
        return []
    # Divide por quebra de linha, ponto e vírgula ou vírgula
    items = re.split(r'\n|;|[•\-\*]\s+', text)
    docs = [i.strip() for i in items if i.strip() and len(i.strip()) > 3]
    return docs if docs else [text.strip()]

def populate_structured_services():
    print("==========================================================")
    print("   POPULANDO TABELAS ESTRUTURADAS DA CARTA DE SERVIÇOS     ")
    print("==========================================================")

    if not os.path.exists(XLSX_PATH):
        # Tenta na pasta CRIADO se não estiver no raiz do bancoia
        alternative_path = os.path.join(ROOT, "data", "knowledge", "CRIADO", "CARTA_DE_SERVICO_AJUSTE_23.05.26.xlsx")
        if os.path.exists(alternative_path):
            path = alternative_path
        else:
            print(f"[Erro] Arquivo de Carta de Serviços não encontrado em {XLSX_PATH}")
            return
    else:
        path = XLSX_PATH

    print(f"Lendo dados estruturados de: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Servicos"]

    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(3, col).value
        headers.append(str(val).strip() if val else f"col_{col}")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Habilita suporte a chaves estrangeiras
    cursor.execute("PRAGMA foreign_keys = ON;")

    # Limpeza seletiva das tabelas estruturadas antes de popular para evitar duplicatas
    cursor.execute("DELETE FROM service_categories;")
    cursor.execute("DELETE FROM service_documents;")
    cursor.execute("DELETE FROM service_steps;")
    cursor.execute("DELETE FROM service_phones;")
    cursor.execute("DELETE FROM service_emails;")
    cursor.execute("DELETE FROM service_links;")
    cursor.execute("DELETE FROM service_priorities;")
    cursor.execute("DELETE FROM services;")
    cursor.execute("DELETE FROM categories;")
    cursor.execute("DELETE FROM secretarias;")
    conn.commit()

    print("Banco estruturado limpo. Iniciando inserção...")

    count_sec = 0
    count_cat = 0
    count_services = 0

    # Cache local de IDs para evitar múltiplas queries
    sec_cache = {} # name -> id
    cat_cache = {} # (sec_id, name) -> id

    for row in range(4, ws.max_row + 1):
        row_data = {}
        vazio = True
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row, col_idx).value
            if val:
                row_data[header] = str(val).strip()
                vazio = False
        
        if vazio:
            continue

        orgao_nome = row_data.get("Órgão", row_data.get("Orgao", "Prefeitura Municipal"))
        cat_nome = row_data.get("Categoria", "Geral")
        servico_nome = row_data.get("Serviço", row_data.get("Servico", ""))

        if not servico_nome:
            continue

        # 1. Garante Secretaria
        if orgao_nome not in sec_cache:
            sec_code = generate_sec_code(orgao_nome)
            # Verifica se código já existe no banco
            cursor.execute("SELECT id FROM secretarias WHERE code = ?", (sec_code,))
            row_sec = cursor.fetchone()
            if row_sec:
                sec_id = row_sec[0]
            else:
                cursor.execute("INSERT INTO secretarias (name, code) VALUES (?, ?)", (orgao_nome, sec_code))
                sec_id = cursor.lastrowid
                count_sec += 1
            sec_cache[orgao_nome] = sec_id
        else:
            sec_id = sec_cache[orgao_nome]

        # 2. Garante Categoria
        cat_key = (sec_id, cat_nome)
        if cat_key not in cat_cache:
            cursor.execute("SELECT id FROM categories WHERE name = ? AND secretaria_id = ?", (cat_nome, sec_id))
            row_cat = cursor.fetchone()
            if row_cat:
                cat_id = row_cat[0]
            else:
                cursor.execute("INSERT INTO categories (name, secretaria_id) VALUES (?, ?)", (cat_nome, sec_id))
                cat_id = cursor.lastrowid
                count_cat += 1
            cat_cache[cat_key] = cat_id
        else:
            cat_id = cat_cache[cat_key]

        # 3. Insere o Serviço
        desc = row_data.get("O que é o serviço", row_data.get("O que e o servico", ""))
        how = row_data.get("Como acessar", "")
        addr = row_data.get("Endereço", row_data.get("Endereco", ""))
        who = row_data.get("Quem pode solicitar", "Cidadão")
        deadline = row_data.get("Prazo", "Não especificado")
        cost = row_data.get("Gratuito", "Gratuito")

        cursor.execute("""
            INSERT INTO services (
                secretaria_id, category_id, name, description, how_to_access, 
                address, who_can_request, max_deadline, cost, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'published')
        """, (sec_id, cat_id, servico_nome, desc, how, addr, who, deadline, cost))
        
        service_id = cursor.lastrowid
        count_services += 1

        # 4. Extração e inserção de telefones, e-mails, links das colunas de texto
        # Procuramos tanto na coluna de canais como no endereço/como acessar
        canal_texto = row_data.get("Canal de atendimento", "")
        context_texto = f"{canal_texto} {how} {addr} {desc}"

        phones = extract_phones(context_texto)
        for ph in set(phones):
            cursor.execute("INSERT INTO service_phones (service_id, phone) VALUES (?, ?)", (service_id, ph))

        emails = extract_emails(context_texto)
        for em in set(emails):
            cursor.execute("INSERT INTO service_emails (service_id, email) VALUES (?, ?)", (service_id, em))

        links = extract_links(context_texto)
        for lk in set(links):
            cursor.execute("INSERT INTO service_links (service_id, link) VALUES (?, ?)", (service_id, lk))

        # 5. Passo a passo (steps) do serviço
        steps = split_steps(how)
        for idx_step, step_desc in enumerate(steps, 1):
            cursor.execute("INSERT INTO service_steps (service_id, step_number, description) VALUES (?, ?, ?)",
                           (service_id, idx_step, step_desc))

        # 6. Documentos necessários
        docs = split_documents(row_data.get("Documentos necessários", row_data.get("Documentos necessarios", "")))
        for doc_name in docs:
            cursor.execute("INSERT INTO service_documents (service_id, document_name) VALUES (?, ?)",
                           (service_id, doc_name))

    conn.commit()
    conn.close()

    print("\n[Sucesso] Processamento da planilha estruturada finalizado!")
    print(f"  Secretarias cadastradas : {count_sec}")
    print(f"  Categorias criadas       : {count_cat}")
    print(f"  Serviços cadastrados    : {count_services}")
    print("==========================================================")

if __name__ == "__main__":
    populate_structured_services()
